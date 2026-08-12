from io import BytesIO

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)


def create_excel_response(
    *,
    filename: str,
    sheet_name: str,
    headers: list[str],
    rows,
    number_columns: set[int] | None = None,
) -> HttpResponse:
    """
    Create an Excel (.xlsx) HTTP response.

    Args:
        filename:
            Output filename.

        sheet_name:
            Excel worksheet name.

        headers:
            Column headers.

        rows:
            Iterable of row values.

        number_columns:
            Zero-based indexes of numeric columns.
    """

    number_columns = number_columns or set()

    # ---------------------------------------------------------
    # Workbook
    # ---------------------------------------------------------

    workbook = Workbook()

    # Remove the default worksheet.
    default_worksheet = workbook.active

    if default_worksheet is not None:
        workbook.remove(default_worksheet)

    # Create a new worksheet.
    worksheet = workbook.create_sheet(
        title=sheet_name[:31],
    )

    # ---------------------------------------------------------
    # Direction
    # ---------------------------------------------------------

    worksheet.sheet_view.rightToLeft = True

    # ---------------------------------------------------------
    # Header
    # ---------------------------------------------------------

    worksheet.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F2937",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    thin_border = Border(
        bottom=Side(
            style="thin",
            color="D1D5DB",
        )
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # ---------------------------------------------------------
    # Data
    # ---------------------------------------------------------

    for row in rows:
        worksheet.append(list(row))

    # ---------------------------------------------------------
    # Cell formatting
    # ---------------------------------------------------------

    if worksheet.max_row >= 2:
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for index, cell in enumerate(row):
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="right",
                )

                if index in number_columns:
                    cell.number_format = "#,##0"

    # ---------------------------------------------------------
    # Freeze header
    # ---------------------------------------------------------

    worksheet.freeze_panes = "A2"

    # ---------------------------------------------------------
    # Auto filter / table
    # ---------------------------------------------------------

    if worksheet.max_row >= 2 and worksheet.max_column >= 1:
        last_column = get_column_letter(worksheet.max_column)

        table_ref = f"A1:{last_column}{worksheet.max_row}"

        table = Table(
            displayName="ExportTable",
            ref=table_ref,
        )

        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = table_style

        worksheet.add_table(table)

    # ---------------------------------------------------------
    # Column width
    # ---------------------------------------------------------

    for column_cells in worksheet.columns:
        if not column_cells:
            continue

        max_length = 0

        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            value_length = len(str(value))

            max_length = max(
                max_length,
                value_length,
            )

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            45,
        )

    # ---------------------------------------------------------
    # Row height
    # ---------------------------------------------------------

    worksheet.row_dimensions[1].height = 24

    # ---------------------------------------------------------
    # Response
    # ---------------------------------------------------------

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-" "officedocument.spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response

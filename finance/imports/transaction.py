import csv
import io
import logging
import uuid
from datetime import date, datetime
from typing import Any, BinaryIO
from django.contrib.auth import get_user_model
from django.db import transaction as db_transaction

from finance.enums import CategoryType, TransactionType
from finance.models import CategoryModel, TransactionModel

logger = logging.getLogger("finance.transaction.import")

UserModel = get_user_model()


class TransactionImportService:
    """
    Import transactions from CSV / Excel.

    Import rules:

    - Existing transaction ID => skipped.
    - New transaction ID => created.
    - Category is optional.
    - Missing category => transaction is created without category.
    - Non-existing category => automatically created.
    - Missing category type => inferred from transaction type.
    - Amount is always stored as a positive value.
    - Transaction type determines income / expense.
    - User is resolved by email when provided.
    - If user email is missing, the authenticated admin user is used.
    """

    # =========================================================
    # Public API
    # =========================================================

    @classmethod
    def import_csv(
        cls,
        *,
        user,
        file: BinaryIO,
    ) -> dict[str, Any]:
        """
        Import transactions from CSV.
        """

        content = file.read()

        if isinstance(content, bytes):
            text = content.decode("utf-8-sig")
        else:
            text = content

        reader = csv.DictReader(
            io.StringIO(text),
        )

        if not reader.fieldnames:
            raise ValueError("فایل CSV فاقد ستون است.")

        rows: list[dict[str, Any]] = []

        for raw_row in reader:
            row: dict[str, Any] = {}

            for key, value in raw_row.items():
                normalized_key = cls._normalize_header(key)

                if not normalized_key:
                    continue

                row[normalized_key] = value

            rows.append(row)

        return cls._import_rows(
            user=user,
            rows=rows,
        )

    # =========================================================
    # Excel
    # =========================================================

    @classmethod
    def import_excel(
        cls,
        *,
        user,
        file: BinaryIO,
    ) -> dict[str, Any]:
        """
        Import transactions from XLSX.
        """

        from openpyxl import load_workbook

        workbook = load_workbook(
            file,
            read_only=True,
            data_only=True,
        )

        try:
            worksheet = workbook.active

            if worksheet is None:
                raise ValueError("فایل Excel فاقد worksheet معتبر است.")

            rows_iterator = worksheet.iter_rows(
                values_only=True,
            )

            try:
                raw_headers = next(rows_iterator)
            except StopIteration as exc:
                raise ValueError("فایل Excel خالی است.") from exc

            headers: list[str] = []

            for value in raw_headers:
                normalized_header = cls._normalize_header(value)

                headers.append(normalized_header)

            # -------------------------------------------------
            # Validate that headers actually exist
            # -------------------------------------------------

            if not any(headers):
                raise ValueError("فایل Excel فاقد ستون معتبر است.")

            rows: list[dict[str, Any]] = []

            for raw_row in rows_iterator:
                row: dict[str, Any] = {}

                for index, header in enumerate(headers):
                    if not header:
                        continue

                    value = raw_row[index] if index < len(raw_row) else None

                    row[header] = value

                # Ignore completely empty rows.
                if any(
                    value is not None and str(value).strip() for value in row.values()
                ):
                    rows.append(row)

            return cls._import_rows(
                user=user,
                rows=rows,
            )

        finally:
            workbook.close()

    # =========================================================
    # Main Import
    # =========================================================

    @classmethod
    def _import_rows(
        cls,
        *,
        user,
        rows: list[dict[str, Any]],
    ) -> dict[str, Any]:

        if not rows:
            return {
                "total": 0,
                "created": 0,
                "skipped": 0,
                "failed": 0,
                "errors": [],
            }

        normalized_rows: list[dict[str, Any]] = [
            cls._normalize_row(row) for row in rows
        ]

        cls._validate_headers(normalized_rows)

        result: dict[str, Any] = {
            "total": len(normalized_rows),
            "created": 0,
            "skipped": 0,
            "failed": 0,
            "errors": [],
        }

        for index, row in enumerate(
            normalized_rows,
            start=2,
        ):
            try:
                with db_transaction.atomic():

                    # =================================================
                    # Transaction ID
                    # =================================================

                    transaction_id = cls._parse_uuid(row.get("id"))

                    # =================================================
                    # Existing transaction
                    # =================================================

                    if TransactionModel.objects.filter(
                        id=transaction_id,
                    ).exists():

                        result["skipped"] += 1
                        continue

                    # =================================================
                    # User
                    # =================================================

                    transaction_user = cls._resolve_user(
                        row=row,
                        default_user=user,
                    )

                    # =================================================
                    # Transaction Type
                    # =================================================

                    transaction_type = cls._resolve_transaction_type(
                        row.get("transaction_type")
                    )

                    # =================================================
                    # Category
                    # =================================================

                    category = cls._resolve_category(
                        user=transaction_user,
                        row=row,
                        transaction_type=transaction_type,
                    )

                    # =================================================
                    # Amount
                    # =================================================

                    amount = cls._parse_amount(row.get("amount"))

                    # =================================================
                    # Date
                    # =================================================

                    transaction_date = cls._parse_date(row.get("date"))

                    # =================================================
                    # Description
                    # =================================================

                    description = cls._clean_string(row.get("description"))

                    # =================================================
                    # Create
                    # =================================================

                    create_kwargs: dict[str, Any] = {
                        "id": transaction_id,
                        "user": transaction_user,
                        "category": category,
                        "amount": amount,
                        "type": transaction_type,
                        "description": description,
                    }

                    if transaction_date is not None:
                        create_kwargs["date"] = transaction_date

                    TransactionModel.objects.create(**create_kwargs)

                    result["created"] += 1

            except Exception as exc:
                logger.exception(
                    "Transaction import failed at row %s",
                    index,
                )

                result["failed"] += 1

                result["errors"].append(
                    {
                        "row": index,
                        "message": str(exc),
                    }
                )

        return result

    # =========================================================
    # Header Validation
    # =========================================================

    @classmethod
    def _validate_headers(
        cls,
        rows: list[dict[str, Any]],
    ) -> None:
        """
        Validate only truly required columns.

        Category and category type are optional.
        User email is optional.
        Date is optional.
        Description is optional.
        """

        if not rows:
            return

        available = set(rows[0].keys())

        required = {
            "id",
            "transaction_type",
            "amount",
        }

        missing = required - available

        if not missing:
            return

        labels = {
            "id": "شناسه",
            "transaction_type": "نوع تراکنش",
            "amount": "مبلغ",
        }

        missing_labels = ", ".join(
            labels.get(
                field,
                field,
            )
            for field in sorted(missing)
        )

        raise ValueError(f"ستون‌های الزامی وجود ندارند: " f"{missing_labels}")

    # =========================================================
    # Header Normalization
    # =========================================================

    @classmethod
    def _normalize_header(
        cls,
        value: Any,
    ) -> str:
        """
        Normalize CSV / Excel column names.

        Always returns str.
        """

        if value is None:
            return ""

        normalized = str(value)

        # BOM
        normalized = normalized.replace(
            "\ufeff",
            "",
        )

        normalized = normalized.strip()

        # Persian / Arabic characters
        normalized = normalized.replace(
            "ي",
            "ی",
        )

        normalized = normalized.replace(
            "ك",
            "ک",
        )

        aliases: dict[str, str] = {
            # =================================================
            # ID
            # =================================================
            "شناسه": "id",
            "شناسه تراکنش": "id",
            "id": "id",
            "transaction_id": "id",
            "transaction id": "id",
            # =================================================
            # User
            # =================================================
            "ایمیل کاربر": "user_email",
            "ایمیل": "user_email",
            "user_email": "user_email",
            "user email": "user_email",
            "email": "user_email",
            # =================================================
            # Category
            # =================================================
            "دسته‌بندی": "category_name",
            "دسته بندی": "category_name",
            "دسته": "category_name",
            "category": "category_name",
            "category_name": "category_name",
            "category name": "category_name",
            # =================================================
            # Category Type
            # =================================================
            "نوع دسته‌بندی": "category_type",
            "نوع دسته بندی": "category_type",
            "category_type": "category_type",
            "category type": "category_type",
            # =================================================
            # Transaction Type
            # =================================================
            "نوع تراکنش": "transaction_type",
            "نوع": "transaction_type",
            "transaction_type": "transaction_type",
            "transaction type": "transaction_type",
            "type": "transaction_type",
            # =================================================
            # Amount
            # =================================================
            "مبلغ": "amount",
            "amount": "amount",
            "مبلغ مطلق": "absolute_amount",
            "مبلغ خالص": "absolute_amount",
            "absolute_amount": "absolute_amount",
            "absolute amount": "absolute_amount",
            # =================================================
            # Description
            # =================================================
            "توضیحات": "description",
            "توضیح": "description",
            "description": "description",
            # =================================================
            # Gregorian Date
            # =================================================
            "تاریخ میلادی": "date",
            "تاریخ": "date",
            "date": "date",
            # =================================================
            # Persian Date
            # =================================================
            "تاریخ شمسی": "persian_date",
            "persian_date": "persian_date",
            # =================================================
            # Persian Month
            # =================================================
            "ماه شمسی": "persian_month_name",
            "ماه": "persian_month_name",
            # =================================================
            # Persian Year
            # =================================================
            "سال شمسی": "persian_year",
            "سال": "persian_year",
            # =================================================
            # Month Number
            # =================================================
            "شماره ماه": "month",
            "month": "month",
            # =================================================
            # Created At
            # =================================================
            "تاریخ ایجاد": "created_at",
            "created_at": "created_at",
            "created at": "created_at",
        }

        alias = aliases.get(normalized)

        if alias is not None:
            return alias

        # IMPORTANT:
        # normalized is guaranteed to be str.
        return normalized.lower().replace(
            " ",
            "_",
        )

    # =========================================================
    # Row Normalization
    # =========================================================

    @classmethod
    def _normalize_row(
        cls,
        row: dict[str, Any],
    ) -> dict[str, Any]:

        normalized: dict[str, Any] = {}

        for key, value in row.items():

            normalized_key = cls._normalize_header(key)

            if not normalized_key:
                continue

            normalized[normalized_key] = value

        return normalized

    # =========================================================
    # User
    # =========================================================

    @classmethod
    def _resolve_user(
        cls,
        *,
        row: dict[str, Any],
        default_user,
    ):

        email = cls._clean_string(row.get("user_email"))

        # If exported/imported file doesn't contain
        # user email, use currently authenticated admin user.
        if not email:
            return default_user

        try:
            return UserModel.objects.get(
                email__iexact=email,
            )

        except UserModel.DoesNotExist as exc:
            raise ValueError(f"کاربری با ایمیل «{email}» پیدا نشد.") from exc

    # =========================================================
    # Transaction Type
    # =========================================================

    @classmethod
    def _resolve_transaction_type(
        cls,
        value: Any,
    ) -> TransactionType:

        cleaned = cls._clean_string(value)

        if not cleaned:
            raise ValueError("نوع تراکنش الزامی است.")

        normalized = cleaned.lower()

        mapping: dict[str, TransactionType] = {
            # English
            "income": TransactionType.INCOME,
            "expense": TransactionType.EXPENSE,
            # Persian
            "درآمد": TransactionType.INCOME,
            "هزینه": TransactionType.EXPENSE,
            "دریافتی": TransactionType.INCOME,
            "پرداختی": TransactionType.EXPENSE,
        }

        transaction_type = mapping.get(normalized)

        if transaction_type is None:
            raise ValueError(f"نوع تراکنش «{cleaned}» معتبر نیست.")

        return transaction_type

    # =========================================================
    # Category
    # =========================================================

    @classmethod
    def _resolve_category(
        cls,
        *,
        user,
        row: dict[str, Any],
        transaction_type: TransactionType,
    ) -> CategoryModel | None:

        category_name = cls._clean_string(row.get("category_name"))

        # Category is optional.
        if not category_name:
            return None

        category_type = cls._resolve_category_type(
            row.get("category_type"),
            transaction_type,
        )

        # =====================================================
        # Find existing category
        # =====================================================

        category = CategoryModel.objects.filter(
            user=user,
            name__iexact=category_name,
            type=category_type,
        ).first()

        if category is not None:
            return category

        # =====================================================
        # Create category automatically
        # =====================================================

        return CategoryModel.objects.create(
            user=user,
            name=category_name,
            type=category_type,
        )

    # =========================================================
    # Category Type
    # =========================================================

    @classmethod
    def _resolve_category_type(
        cls,
        value: Any,
        transaction_type: TransactionType,
    ) -> CategoryType:

        cleaned = cls._clean_string(value)

        if cleaned:
            normalized = cleaned.lower()

            mapping: dict[str, CategoryType] = {
                "income": CategoryType.INCOME,
                "expense": CategoryType.EXPENSE,
                "درآمد": CategoryType.INCOME,
                "هزینه": CategoryType.EXPENSE,
            }

            category_type = mapping.get(normalized)

            if category_type is not None:
                return category_type

        # Infer from transaction type.
        if transaction_type == TransactionType.INCOME:
            return CategoryType.INCOME

        return CategoryType.EXPENSE

    # =========================================================
    # UUID
    # =========================================================

    @classmethod
    def _parse_uuid(
        cls,
        value: Any,
    ) -> uuid.UUID:

        cleaned = cls._clean_string(value)

        if not cleaned:
            raise ValueError("شناسه تراکنش الزامی است.")

        try:
            return uuid.UUID(cleaned)

        except (
            ValueError,
            AttributeError,
            TypeError,
        ) as exc:

            raise ValueError("شناسه تراکنش معتبر نیست.") from exc

    # =========================================================
    # Amount
    # =========================================================

    @classmethod
    def _parse_amount(
        cls,
        value: Any,
    ) -> int:

        if value is None:
            raise ValueError("مبلغ الزامی است.")

        if isinstance(value, bool):
            raise ValueError("مبلغ معتبر نیست.")

        if isinstance(value, float):
            amount = int(value)

        elif isinstance(value, int):
            amount = value

        else:
            cleaned = cls._clean_string(value)

            if not cleaned:
                raise ValueError("مبلغ الزامی است.")

            cleaned = cls._persian_digits_to_english(cleaned)

            cleaned = cleaned.replace(",", "").replace("٬", "").replace(" ", "")

            try:
                amount = int(float(cleaned))

            except (
                ValueError,
                TypeError,
            ) as exc:

                raise ValueError("مبلغ معتبر نیست.") from exc

        # Transaction model stores
        # amount as positive value.
        amount = abs(amount)

        if amount <= 0:
            raise ValueError("مبلغ باید بیشتر از صفر باشد.")

        return amount

    # =========================================================
    # Date
    # =========================================================

    @classmethod
    def _parse_date(
        cls,
        value: Any,
    ) -> date | None:

        if value is None:
            return None

        # Excel datetime/date object.
        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        cleaned = cls._clean_string(value)

        if not cleaned:
            return None

        formats = (
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
        )

        for date_format in formats:
            try:
                return datetime.strptime(
                    cleaned,
                    date_format,
                ).date()

            except ValueError:
                continue

        raise ValueError(f"تاریخ «{cleaned}» معتبر نیست.")

    # =========================================================
    # Helpers
    # =========================================================

    @staticmethod
    def _clean_string(
        value: Any,
    ) -> str | None:

        if value is None:
            return None

        cleaned = str(value)

        cleaned = cleaned.replace(
            "\ufeff",
            "",
        )

        cleaned = cleaned.strip()

        return cleaned or None

    @staticmethod
    def _persian_digits_to_english(
        value: str,
    ) -> str:

        translation = str.maketrans(
            "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
            "01234567890123456789",
        )

        return value.translate(translation)
